import openpyxl

#variables for excell sheet
starting_rows = 13
starting_column = 12
max_entries = 17
max_attributes = 26

#marks 
drill_wr_total = 10
drill_pr_total = 80
drill_total = 90

writing_wr_total = 35
writing_pr_total = 60
writing_total = 60

misc_total = 200

special_wr_total = 105
special_pr_total = 45 
special_total = 150

total = 500

pass_ind = 0.45
pass_ttl= 0.5


wb = openpyxl.load_workbook('test.xlsx')
sheet = wb["Sheet1"]
sheet = wb.active


# for value in sheet.iter_rows(min_row=starting_rows, max_row= max_entries, min_col=starting_column, max_col=max_attributes, values_only=True):
#     print(value)

for rows in range(starting_rows, max_entries):

    drill_tr = 0
    writing_tr = 0
    misc_tr = 0
    spl_tr = 0
    grand_total = 0
    count = 1

    for columns in range(starting_column, max_attributes):
        if(count == 1):
            drill_wr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 2):
            drill_pr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 3):
            drill_tr =  int(sheet.cell(row = rows, column = columns).value)
        elif(count == 4):
            writing_wr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 5):
            writing_pr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 6):
            writing_tr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 7):
            misc_tr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 8):
            special_wr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 9):
            special_pr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 10):
            spl_tr = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 13):
            grand_total = int(sheet.cell(row = rows, column = columns).value)
        elif(count == 14):
            if(drill_wr > drill_wr_total * pass_ind) and(drill_pr > drill_pr_total * pass_ind) and (drill_tr > drill_total * pass_ind) and (writing_wr > writing_wr_total * pass_ind ) and (writing_pr > writing_pr_total * pass_ind) and (writing_tr > writing_total * pass_ind) and (misc_tr > misc_total* pass_ind) and (special_wr > special_wr_total * pass_ind) and (special_pr > special_pr_total * pass_ind) and (spl_tr > special_total* pass_ind) and (grand_total > total * pass_ttl):
                if(grand_total > 375):
                    # temp = sheet.cell(row = rows, column = max_attributes)
                    # temp.value = 'A'
                    var = 'Y' + str(rows)
                    sheet[var] = "A"
                if(grand_total > 300) and (grand_total < 374):
                    # temp = sheet.cell(row = rows, column = max_attributes)
                    # temp.value = 'B'
                    var = 'Y' + str(rows)
                    sheet[var] = "B"
                if(grand_total > 250) and (grand_total < 299):
                    # temp = sheet.cell(row = rows, column = max_attributes)
                    # temp.value = 'C'
                    var = 'Y' + str(rows)
                    sheet[var] = "C"
            else:
                # temp = sheet.cell(row = rows, column = max_attributes)
                # temp.value = 'Fail'
                var = 'Y' + str(rows)
                sheet[var] = "Fail"
        else:
            b = 2
        count = count + 1

        # if(columns >= starting_column + 0 & columns < starting_column + 2):
        #     drill_tr = drill_tr + int(sheet.cell(row = rows, column = columns).value)
        
        # if(columns >= starting_column + 3   & columns > starting_column + 5):
        #     writing_tr = writing_tr + int(sheet.cell(row = rows, column = columns).value)

        # if(columns == starting_column + 6 ):
        #     misc_tr = int(sheet.cell(row = rows, column = columns).value)

        # if(columns >= starting_column + 7  & columns > starting_column + 9):
        #     spl_tr = spl_tr + int(sheet.cell(row = rows, column = columns).value)
        
        # if(drill_tr >= drill_total*0.45):
        #     if(writing_tr >= writing_total*0.45):
        #         if(misc_tr > misc_total*0.45):
        #             if(spl_tr > special_total*0.45):
        #                 grand_total = drill_tr + writing_tr + misc_tr + spl_tr
        #                 if(grand_total > total*0.5):
        #                     if(grand_total>375):
        #                         temp = sheet.cell(row = rows, column = max_attributes)
        #                         temp.value = "A"
        #                     elif(grand_total > 300 & grand_total < 374):
        #                         temp = sheet.cell(row = rows, column = max_attributes)
        #                         temp.value = "B"
        #                     elif(grand_total > 250 & grand_total < 299):
        #                         temp = sheet.cell(row = rows, column = max_attributes)
        #                         temp.value = "C"
        #                     else:
        #                         temp = sheet.cell(row = rows, column = max_attributes)
        #                         temp.value = "Fail"
        # else:
        #     temp = sheet.cell(row = rows, column = max_attributes)
        #     temp.value = "Fail"
        # print(sheet.cell(row = rows, column = columns).value)
wb.save('test.xlsx')


    

        
        